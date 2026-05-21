"""Export Fair Forward catalog entries (Africa-tagged) to the open-data-africa catalogue.

Reads the canonical Fair Forward catalog JSON, filters to African countries, and
appends new rows to open-data-africa's datasets/catalog.csv. Existing rows are
never modified. After appending, regenerates the downloadable Excel mirror from
the updated CSV and removes the older Open Datasets Africa.xlsx if present.

Usage:
    python scripts/export_to_open_data_africa.py            # apply
    python scripts/export_to_open_data_africa.py --dry-run  # report only

Defaults can be overridden via CLI flags.
"""

import argparse
import csv
import json
import os
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


HERE = Path(__file__).resolve().parent
FF_REPO_ROOT = HERE.parent
DEFAULT_FF_CATALOG = FF_REPO_ROOT / "public" / "data" / "catalog.json"
DEFAULT_ODA_REPO = Path("/Users/jonas/github/open-data-africa.github.io")
DEFAULT_ODA_CSV = DEFAULT_ODA_REPO / "datasets" / "catalog.csv"
DEFAULT_ODA_XLSX_NEW = DEFAULT_ODA_REPO / "datasets" / "Open Datasets and Use Cases Africa.xlsx"
DEFAULT_ODA_XLSX_OLD = DEFAULT_ODA_REPO / "datasets" / "Open Datasets Africa.xlsx"

ODA_COLUMNS = [
    "Dataset Name",
    "Domain",
    "Source",
    "Country",
    "Data Type",
    "License",
    "Link",
    "Use Cases",
    "Last Updated",
    "Challenges/Notes",
]

AFRICAN_COUNTRIES_RAW = [
    "Angola", "Benin", "Burkina Faso", "Burundi", "Cameroon", "Cape Verde",
    "Central African Republic", "Chad", "Comoros", "Congo",
    "Democratic Republic of the Congo", "DRC", "Djibouti", "Egypt",
    "Equatorial Guinea", "Eritrea", "Eswatini", "Ethiopia", "Gabon", "Gambia",
    "Ghana", "Guinea", "Guinea-Bissau", "Ivory Coast", "Côte d'Ivoire",
    "Cote d'Ivoire", "Kenya", "Lesotho", "Liberia", "Libya", "Madagascar",
    "Malawi", "Mali", "Mauritania", "Mauritius", "Morocco", "Mozambique",
    "Namibia", "Niger", "Nigeria", "Rwanda", "São Tomé and Príncipe",
    "Sao Tome and Principe", "Senegal", "Seychelles", "Sierra Leone", "Somalia",
    "South Africa", "South Sudan", "Sudan", "Tanzania", "Togo", "Tunisia",
    "Uganda", "Zambia", "Zimbabwe",
]

SHALLOW_PORTAL_DOMAINS = {
    "gbif.org", "data.gov.gh", "masakhane.io", "arxiv.org", "wikipedia.org",
    "creativecommons.org", "youtube.com", "medium.com", "twitter.com", "x.com",
    "linkedin.com", "facebook.com", "data.world",
}

STOPWORDS = set((
    "the a an and or for from with into onto over under that this these those "
    "new old data dataset datasets set sets project projects use using used "
    "cases case study studies of in on to by at via per how what when where"
).split())


def strip_accents(s):
    if not s:
        return ""
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def norm_text(s):
    s = strip_accents(s or "").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return s


def sig_tokens(s):
    return {t for t in norm_text(s).split() if len(t) > 2 and t not in STOPWORDS}


def norm_url(u):
    if not u:
        return ""
    u = str(u).strip().lower().rstrip("/")
    u = re.sub(r"#.*$", "", u)
    u = re.sub(r"\?utm_[^&]+(&|$)", "", u).rstrip("?&")
    u = re.sub(r"^https?://", "", u)
    u = re.sub(r"^www\.", "", u)
    return u


def url_is_specific(nu):
    """A URL counts as a strong match anchor only if it goes deeper than a
    portal homepage. Portal-only URLs (gbif.org, github.com root, etc.) are
    classed as 'soft' — we flag them for human review rather than auto-matching.
    """
    if not nu:
        return False
    parts = nu.split("/", 1)
    domain = parts[0]
    path = parts[1] if len(parts) > 1 else ""
    if not path:
        return False
    if domain in SHALLOW_PORTAL_DOMAINS:
        if "persistentid" in path:
            return True
        segments = [s for s in path.split("/") if s]
        return len(segments) >= 3 and len(path) >= 25
    # Domains known to be data hosts but commonly with shallow URLs
    if domain == "huggingface.co":
        return len([s for s in path.split("/") if s]) >= 3
    if domain == "github.com":
        return len([s for s in path.split("/") if s]) >= 2
    if domain == "kaggle.com":
        return len([s for s in path.split("/") if s]) >= 2
    if domain == "dataverse.harvard.edu":
        return "persistentid" in path
    if domain == "zenodo.org":
        return "record" in path
    if domain == "figshare.com":
        return "articles" in path
    return True


AFRICA_NORM = {norm_text(c) for c in AFRICAN_COUNTRIES_RAW}


def all_project_urls(project):
    urls = []
    for key in ("dataset_links", "usecase_links", "additional_resources", "hosted_documents"):
        for link in (project.get(key) or []):
            u = link.get("url") if isinstance(link, dict) else str(link)
            if u:
                urls.append(norm_url(u))
    return urls


def derive_domain(project):
    """Map FF data types + SDGs + title keywords to an ODA Domain value."""
    data_types = " ".join(project.get("data_types") or []).lower()
    sdgs = " ".join(project.get("sdgs") or []).lower()
    title = (project.get("title") or "").lower()

    def contains_any(haystack, needles):
        return any(n in haystack for n in needles)

    if contains_any(data_types, ["speech", "audio", "text"]) or contains_any(
        title, ["language", "voice", "speech", "translat", "chatbot", "chat ", "nlp", "ivr"]
    ):
        return "NLP"
    if "sdg 2" in sdgs or contains_any(
        title, ["crop", "farm", "agro", "cocoa", "cashew", "livestock", "farmer", "agrivolt", "harvest"]
    ):
        return "Agriculture"
    if "sdg 3" in sdgs or contains_any(title, ["health", "disease", "medical", "clinic"]):
        return "Healthcare"
    if "sdg 7" in sdgs or contains_any(title, ["solar", "electrif", "energy"]):
        return "Energy"
    if "sdg 11" in sdgs or contains_any(title, ["mobility", "transport", "urban"]):
        return "Mobility"
    if (
        any(t in sdgs for t in ("sdg 13", "sdg 14", "sdg 15"))
        or "geospatial" in data_types
        or contains_any(
            title,
            ["forest", "climate", "mangrove", "biomass", "carbon", "ecolog", "biodivers", "tree"],
        )
    ):
        return "Environment"
    return "Other"


def derive_source(project):
    return (
        (project.get("organizations") or "").strip()
        or (project.get("contact") or "").strip()
        or (project.get("editor") or "").strip()
        or "Fair Forward"
    )


def derive_country_cell(project):
    african = [c for c in (project.get("countries") or []) if norm_text(c) in AFRICA_NORM]
    return "; ".join(african)


def derive_data_type(project):
    return "; ".join(project.get("data_types") or [])


def derive_link(project):
    for key in ("dataset_links", "usecase_links"):
        for link in (project.get(key) or []):
            url = link.get("url") if isinstance(link, dict) else None
            if url:
                return url
    return ""


def derive_use_cases(project):
    desc = (project.get("description") or "").strip()
    if len(desc) > 280:
        desc = desc[:277].rstrip() + "…"
    return desc


def derive_notes(project):
    parts = ["Fair Forward portfolio."]
    maturity = (project.get("maturity") or "").strip()
    if maturity:
        parts.append(maturity)
    if project.get("is_lacuna"):
        parts.append("Lacuna Fund data.")
    return " ".join(parts)


def project_to_row(project):
    return {
        "Dataset Name": project.get("title") or "",
        "Domain": derive_domain(project),
        "Source": derive_source(project),
        "Country": derive_country_cell(project),
        "Data Type": derive_data_type(project),
        "License": project.get("license") or "",
        "Link": derive_link(project),
        "Use Cases": derive_use_cases(project),
        "Last Updated": "2026",
        "Challenges/Notes": derive_notes(project),
    }


def load_existing_rows(csv_path):
    with csv_path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def build_match_index(existing_rows):
    specific_urls = set()
    soft_urls = set()
    titles = []
    for row in existing_rows:
        link = (row.get("Link") or "").strip()
        nu = norm_url(link)
        if nu:
            (specific_urls if url_is_specific(nu) else soft_urls).add(nu)
        title = (row.get("Dataset Name") or "").strip()
        if title:
            titles.append(title)
    return specific_urls, soft_urls, titles


def classify_project(project, specific_urls, soft_urls, titles):
    """Return one of: 'present', 'soft-collision', 'new', plus evidence string."""
    ff_urls = [u for u in all_project_urls(project) if u]
    for u in ff_urls:
        if url_is_specific(u) and u in specific_urls:
            return "present", f"specific URL: {u}"

    ft = sig_tokens(project.get("title") or "")
    nft = norm_text(project.get("title") or "")
    if len(ft) >= 3:
        for t in titles:
            ott = sig_tokens(t)
            if not ott:
                continue
            inter = ft & ott
            union = ft | ott
            if union and len(inter) / len(union) >= 0.65 and len(inter) >= 3:
                return "present", f"title jaccard {len(inter)}/{len(union)} vs '{t[:60]}'"
    if len(nft) >= 35:
        for t in titles:
            nor = norm_text(t)
            if len(nor) >= 35 and (nft in nor or nor in nft):
                return "present", f"title containment vs '{t[:60]}'"

    for u in ff_urls:
        if u in soft_urls or (u in specific_urls and not url_is_specific(u)):
            return "soft-collision", f"shallow URL match: {u}"

    return "new", ""


def append_csv(csv_path, new_rows):
    """Append new rows to the CSV without rewriting existing bytes."""
    if not new_rows:
        return
    needs_newline = False
    with csv_path.open("rb") as f:
        try:
            f.seek(-1, os.SEEK_END)
            last = f.read(1)
            needs_newline = last not in (b"\n", b"\r")
        except OSError:
            needs_newline = False
    with csv_path.open("a", encoding="utf-8", newline="") as f:
        if needs_newline:
            f.write("\r\n")
        writer = csv.DictWriter(
            f,
            fieldnames=ODA_COLUMNS,
            quoting=csv.QUOTE_MINIMAL,
            lineterminator="\r\n",
        )
        for row in new_rows:
            writer.writerow({k: row.get(k, "") for k in ODA_COLUMNS})


def regenerate_xlsx(csv_path, xlsx_path):
    """Build a single-sheet workbook that mirrors the CSV 1:1."""
    rows = load_existing_rows(csv_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Open Data Africa"
    ws.append(ODA_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(col, "") for col in ODA_COLUMNS])
    ws.freeze_panes = "A2"
    for col_idx, col_name in enumerate(ODA_COLUMNS, start=1):
        max_len = len(col_name)
        for row in rows:
            v = row.get(col_name) or ""
            if v and len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)
    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--ff-catalog", type=Path, default=DEFAULT_FF_CATALOG)
    parser.add_argument("--oda-csv", type=Path, default=DEFAULT_ODA_CSV)
    parser.add_argument("--oda-xlsx-new", type=Path, default=DEFAULT_ODA_XLSX_NEW)
    parser.add_argument("--oda-xlsx-old", type=Path, default=DEFAULT_ODA_XLSX_OLD)
    parser.add_argument("--dry-run", action="store_true", help="Report only; do not write any files.")
    args = parser.parse_args()

    if not args.ff_catalog.exists():
        sys.exit(f"FF catalog not found: {args.ff_catalog}")
    if not args.oda_csv.exists():
        sys.exit(f"ODA catalog.csv not found: {args.oda_csv}")

    catalog = json.loads(args.ff_catalog.read_text(encoding="utf-8"))
    projects = catalog.get("projects") or []
    africa_projects = [
        p for p in projects
        if {norm_text(c) for c in (p.get("countries") or [])} & AFRICA_NORM
    ]

    existing_rows = load_existing_rows(args.oda_csv)
    specific_urls, soft_urls, titles = build_match_index(existing_rows)

    new_rows, present, soft_collision = [], [], []
    for p in africa_projects:
        verdict, evidence = classify_project(p, specific_urls, soft_urls, titles)
        if verdict == "new":
            new_rows.append(project_to_row(p))
        elif verdict == "present":
            present.append((p.get("title") or "", evidence))
        else:
            soft_collision.append((p.get("title") or "", evidence))

    print(f"FF projects total                : {len(projects)}")
    print(f"FF Africa-tagged projects        : {len(africa_projects)}")
    print(f"Already present in ODA (skip)    : {len(present)}")
    print(f"Soft URL collision (skip, review): {len(soft_collision)}")
    print(f"New rows to append               : {len(new_rows)}")
    print(f"Existing ODA CSV rows            : {len(existing_rows)}")
    print(f"Post-append CSV row count        : {len(existing_rows) + len(new_rows)}")

    if soft_collision:
        print("\nSoft-collision projects (NOT added, manual review):")
        for title, ev in soft_collision:
            print(f"  - {title[:90]}  ({ev})")

    if present:
        print("\nProjects already present (NOT modified):")
        for title, ev in present:
            print(f"  - {title[:90]}  ({ev})")

    if args.dry_run:
        print("\n[dry-run] No files written.")
        return

    append_csv(args.oda_csv, new_rows)
    regenerate_xlsx(args.oda_csv, args.oda_xlsx_new)
    if args.oda_xlsx_old.exists() and args.oda_xlsx_old != args.oda_xlsx_new:
        args.oda_xlsx_old.unlink()

    final_rows = load_existing_rows(args.oda_csv)
    print(f"\nDone.")
    print(f"  Wrote {len(new_rows)} new rows to {args.oda_csv}")
    print(f"  Regenerated {args.oda_xlsx_new} (data rows: {len(final_rows)})")
    if args.oda_xlsx_old != args.oda_xlsx_new:
        print(f"  Removed {args.oda_xlsx_old}")


if __name__ == "__main__":
    main()
